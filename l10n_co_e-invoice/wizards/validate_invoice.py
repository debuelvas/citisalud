# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import logging
import io
from datetime import datetime, date
from dateutil.parser import parse
import pytz
import re
from dateutil.parser import parse as parse_date
_logger = logging.getLogger(__name__)

try:
    import xlrd
    import openpyxl
except ImportError:
    _logger.error("No se puede importar xlrd y/o openpyxl")

_logger = logging.getLogger(__name__)
class ValidateInvoice(models.TransientModel):
    _name = 'ati.validate.invoice'
    _description = "Wizard - Validate multiple invoice"

    def validate_invoice(self):
        """
        Procesa facturas en lotes de 10, validando su estado antes de publicar
        """
        BATCH_SIZE = 10
        invoices = self._context.get('active_ids', [])
        total_invoices = len(invoices)
        processed = 0
        failed = []
        
        try:
            for start in range(0, total_invoices, BATCH_SIZE):
                batch = invoices[start:start + BATCH_SIZE]
                
                for invoice_id in batch:
                    try:
                        invoice = self.env['account.move'].browse(invoice_id)
                        
                        if not invoice.exists():
                            failed.append((invoice_id, _("Factura no encontrada")))
                            continue
                        
                        if invoice.state == 'posted':
                            _logger.info(f'Factura {invoice.name} ya está publicada, procediendo con validación DIAN')
                            invoice.validate_dian()
                        elif invoice.state == 'draft':
                            _logger.info(f'Publicando factura {invoice.name} y validando en DIAN')
                            invoice.action_post()
                            invoice.validate_dian()
                        else:
                            failed.append((invoice.name, f"Estado inválido: {invoice.state}"))
                            continue
                            
                        processed += 1
                        
                        if len(batch) == BATCH_SIZE:
                            self.env.cr.commit()
                    except Exception as e:
                        failed.append((invoice.name if invoice else invoice_id, str(e)))
                        _logger.error(f"Error procesando factura {invoice.name if invoice else invoice_id}: {str(e)}")
                        continue
                        
                _logger.info(f'Procesadas {processed} de {total_invoices} facturas')
        except Exception as e:
            raise UserError(_("Error en el proceso por lotes: %s") % str(e))
        
        finally:
            message = f"Proceso completado:\n"
            message += f"- Facturas procesadas: {processed}/{total_invoices}\n"
            
            if failed:
                message += "\nFacturas con error:\n"
                for invoice_name, error in failed:
                    message += f"- {invoice_name}: {error}\n"
                    
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Resultado de la Validación'),
                    'message': message,
                    'type': 'info' if not failed else 'warning',
                    'sticky': True,
                }
            }

    def _validate_batch(self, batch):
        """
        Valida un lote de facturas
        Returns:
            tuple: (procesadas, fallidas)
        """
        processed = []
        failed = []
        
        for invoice_id in batch:
            try:
                invoice = self.env['account.move'].browse(invoice_id)
                if not invoice.exists():
                    failed.append((invoice_id, "Factura no encontrada"))
                    continue
                    
                if invoice.state == 'posted':
                    invoice.validate_dian()
                elif invoice.state == 'draft':
                    invoice.action_post()
                    invoice.validate_dian()
                else:
                    failed.append((invoice.name, f"Estado inválido: {invoice.state}"))
                    continue
                    
                processed.append(invoice.name)
                
            except Exception as e:
                failed.append((invoice.name if invoice else invoice_id, str(e)))
                
        return processed, failed

class AccountMoveReversal(models.TransientModel):
    _inherit = "account.move.reversal"

    concepto_credit_note = fields.Selection(
        [("1", "Devolución parcial de los bienes y/o no aceptación parcial del servicio"),
        ("2", "Anulación de factura electrónica"),
        ("3", "Rebaja total aplicada"),
        ("4", "Ajuste de precio"),
        ("5", "Descuento comercial por pronto pago"),
        ("6", "Descuento comercial por volumen de ventas")],
        string="Concepto Corrección",
    )
    concept_debit_note = fields.Selection(
        [
            ("1", "Intereses"),
            ("2", "Gastos por cobrar"),
            ("3", "Cambio del valor"),
            ("4", "Otros"),
        ],
        u"Debito Concepto Corrección",
    )
    def reverse_moves(self, is_modify=False):
        res = super().reverse_moves(is_modify=is_modify)
        credit_note = self.env["account.move"].browse(res["res_id"])
        moves_old = self.move_ids
        for rec in moves_old:
            credit_note.write({ "concept_debit_note": self.concept_debit_note,
                            "concepto_credit_note": self.concepto_credit_note,
                            })
        return res
class ValidateInvoice(models.TransientModel):
    _name = 'application.response.wizard'
    _description = "Wizard - Eventos Dian"



class ImportInvoiceWizard(models.TransientModel):
    _name = 'health.invoice.import.wizard'
    _description = 'Asistente para importar facturas del sector salud'

    file = fields.Binary(string='Archivo Excel', required=True)
    file_name = fields.Char(string='Nombre del archivo')
    
    journal_id = fields.Many2one(
        'account.journal', 
        string='Diario por defecto', 
        domain=[('type', '=', 'sale')],
        help='Se usará solo si no se especifica en el archivo o no se encuentra'
    )
    
    invoice_date = fields.Date(
        string='Fecha de factura por defecto',
        default=fields.Date.context_today,
        help='Se usará solo si no se especifica en el archivo o hay errores en el formato'
    )
    
    company_id = fields.Many2one(
        'res.company', 
        string='Compañía',
        default=lambda self: self.env.company,
        required=True
    )
    
    create_missing_partners = fields.Boolean(
        string='Crear clientes no encontrados', 
        default=False,
        help='Crea automáticamente los clientes que no existan en el sistema'
    )
    
    create_missing_products = fields.Boolean(
        string='Crear productos no encontrados', 
        default=False,
        help='Crea automáticamente los productos que no existan en el sistema'
    )
    
    create_missing_patients = fields.Boolean(
        string='Crear pacientes no encontrados', 
        default=False,
        help='Crea automáticamente los pacientes que no existan en el sistema'
    )
    
    default_account_id = fields.Many2one(
        'account.account', 
        string='Cuenta contable por defecto',
        domain="[('deprecated', '=', False), ('company_id', '=', company_id)]",
        help='Cuenta contable a usar para productos creados automáticamente'
    )
    
    # Campos de resultado
    error_log = fields.Binary('Archivo de errores', readonly=True)
    error_filename = fields.Char('Nombre archivo de errores', readonly=True)
    has_errors = fields.Boolean('Tiene errores', default=False)
    error_count = fields.Integer('Cantidad de errores', default=0)
    
    # Estadísticas
    total_invoices_processed = fields.Integer('Facturas procesadas', readonly=True)
    total_lines_processed = fields.Integer('Líneas procesadas', readonly=True)
    total_invoices_failed = fields.Integer('Facturas fallidas', readonly=True)
    total_lines_failed = fields.Integer('Líneas fallidas', readonly=True)
    
    def _get_field_mapping(self):
        """Mapeo exacto de los campos del Excel"""
        return {
            'ID Factura Temporal': 'id_temp',
            'Cliente NIT/ID': 'partner_id',
            'Fecha Factura': 'invoice_date',
            'Fecha Vencimiento': 'invoice_date_due',
            'Diario Código': 'journal_id',
            'Numero Ingreso': 'x_ingreso',
            'Términos de Pago Nombre': 'invoice_payment_term_id',
            'Tipo Operación FE Código': 'invoice_type',
            'Tipo Factura Salud Código': 'fe_operation_type',
            'Fecha Inicio Periodo Servicio': 'date_start',
            'Fecha Fin Periodo Servicio': 'date_end',
            'Modalidad Pago Salud Código': 'health_payment_mode_id',
            'Cobertura Salud Código': 'health_coverage_plan_id',
            'Factura Original Recaudo Número': 'original_invoice_id',
            'Número MIPRES General': 'health_mipres_number',
            'Número Entrega MIPRES General': 'health_mipres_delivery_number',
            
            # Campos de línea
            'Producto/Servicio Código/Nombre': 'product_code',
            'Descripción Detallada Línea': 'name',
            'Diagnostico': 'diagnostico_principal',
            'Paciente ID/Cédula - Línea': 'patient_document',
            'Paciente/nombre': 'patient_name',
            'Ciudad': 'patient_city_id',
            'Nacialidad': 'patient_nationality',
            'Pais': 'patient_country_id',
            'Genero': 'patient_gender',
            'Tipo de usuario': 'patient_user_type',
            'Zona': 'patient_zone',
            'Fecha de nacimineto': 'patient_birth_date',
            'Tipo de documento paciente': 'patient_doc_type',
            'Número Autorización Línea': 'autorizacion',
            'Contrato Nombre/ID - Línea': 'contract_id',
            'Cantidad': 'quantity',
            'Precio Unitario': 'price_unit',
            'Impuestos Nombres': 'tax_ids',
            'Cuenta Contable Código': 'account_id',
            'Fecha Atención Línea': 'fecha_atencion',
        }
    
    def _is_excel_error(self, value):
        """Detecta valores de error de Excel"""
        if not value:
            return False
        error_values = ['#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!']
        return str(value).upper() in error_values
    
    def _validate_date(self, value, field_name, row_num):
        """Valida y convierte fechas con manejo de errores."""
        if not value:
            return False, None

        if self._is_excel_error(value):
            return False, f"Fila {row_num}: Campo '{field_name}' contiene error de Excel: {value}"

        if isinstance(value, date):
            return (value.date() if isinstance(value, datetime) else value), None

        # 4) Fecha serial de Excel (número)
        if isinstance(value, (int, float)):
            try:
                import xlrd
                dt = xlrd.xldate.xldate_as_datetime(value, 0)
                return dt.date(), None
            except Exception as e:
                return False, f"Fila {row_num}: Error al convertir fecha Excel en '{field_name}': {e}"

        s = str(value).strip()
        formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']
        for fmt in formatos:
            try:
                return datetime.strptime(s, fmt).date(), None
            except ValueError:
                pass

        try:
            return parse_date(s, dayfirst=True).date(), None
        except Exception:
            return False, f"Fila {row_num}: Formato de fecha inválido en '{field_name}': {s}"
    def _validate_float(self, value, field_name, row_num, required=False):
        """Valida y convierte valores float con manejo de errores"""
        if not value and not required:
            return 0.0, None
            
        if self._is_excel_error(value):
            return 0.0, f"Fila {row_num}: Campo '{field_name}' contiene error de Excel: {value}"
            
        try:
            if isinstance(value, (int, float)):
                return float(value), None
            return float(str(value).replace(',', '')), None
        except:
            if required:
                return 0.0, f"Fila {row_num}: Valor numérico inválido en '{field_name}': {value}"
            return 0.0, None
    
    def _find_partner(self, partner_nit):
        """Busca cliente por NIT"""
        if not partner_nit:
            return False
            
        partner_nit = str(partner_nit).strip()
        
        partner = self.env['res.partner'].search([
            '|', ('vat', '=', partner_nit), ('vat_co', '=', partner_nit)
        ], limit=1)
        
        if not partner and self.create_missing_partners:
            partner = self.env['res.partner'].create({
                'name': f'Cliente {partner_nit}',
                'vat': partner_nit,
                'company_id': self.company_id.id,
                'company_type': 'company',
            })
            
        return partner
    
    def _find_product(self, product_code):
        """Busca producto por código"""
        if not product_code:
            return False
            
        product_code = str(product_code).strip()
        
        product = self.env['product.product'].search([
            ('default_code', '=', product_code)
        ], limit=1)
        
        if not product and self.create_missing_products:
            product = self.env['product.product'].create({
                'name': f'Producto {product_code}',
                'default_code': product_code,
                'type': 'service',
                'invoice_policy': 'order',
                'property_account_income_id': self.default_account_id.id if self.default_account_id else False,
            })
        
        return product
    
    def _find_or_create_patient(self, line_data):
        """Busca o crea paciente"""
        patient_document = line_data.get('patient_document')
        if not patient_document:
            return False
        
        patient_document = str(patient_document).strip()
        
        patient = self.env['hms.patient'].search([
            '|', ('vat', '=', patient_document), ('ref', '=', patient_document)
        ], limit=1)
        
        if not patient and self.create_missing_patients:
            patient_vals = {
                'name': line_data.get('patient_name') or f'Paciente {patient_document}',
                'vat': patient_document,
                'ref': patient_document,
                'patient': True,
            }
            
            # Fecha de nacimiento
            if line_data.get('patient_birth_date'):
                birth_date, _ = self._validate_date(line_data.get('patient_birth_date'), 'patient_birth_date', 0)
                if birth_date:
                    patient_vals['date_of_birth'] = birth_date
            
            patient = self.env['hms.patient'].create(patient_vals)
        
        return patient
    
    def _find_journal(self, journal_code):
        """Busca diario por código"""
        if not journal_code:
            return self.journal_id
            
        journal = self.env['account.journal'].search([
            ('code', '=', journal_code), 
            ('type', '=', 'sale'),
            ('company_id', '=', self.company_id.id)
        ], limit=1)
        
        return journal or self.journal_id
    
    def _find_payment_term(self, term_name):
        """Busca términos de pago"""
        if not term_name:
            return False
        
        payment_term = self.env['account.payment.term'].search([
            ('name', 'ilike', str(term_name)),
            '|', ('company_id', '=', self.company_id.id), ('company_id', '=', False)
        ], limit=1)
        
        return payment_term
    
    def _find_health_payment_mode(self, code):
        """Busca modalidad de pago por código"""
        if not code:
            return False

        return self.env['health.payment.mode'].search([
            ('code', '=', str(code).strip())
        ], limit=1)
    
    def _find_health_collection_concept(self, code):
        """Busca concepto de recaudo por código"""
        if not code:
            return False
        
        return self.env['health.collection.concept'].search([
            ('code', '=', str(code).strip())
        ], limit=1)
    
    def _find_health_coverage_plan(self, code):
        """Busca cobertura por código"""
        if not code:
            return False
        
        return self.env['health.coverage.plan'].search([
            ('code', '=', str(code).strip())
        ], limit=1)
    
    def _find_contract(self, contract_code):
        """Busca contrato por código o nombre"""
        if not contract_code:
            return False
        
        return self.env['customer.contract'].search([
            ('name', '=', str(contract_code).strip())
        ], limit=1)
    
    def _map_fe_operation_type(self, value):
        """Mapea el tipo de operación FE"""
        if not value:
            return 'SS-CUFE'
        
        value = str(value).strip()
        
        fe_mapping = {
            'ASEGURADORA': 'SS-CUFE',
            'SS-CUFE': 'SS-CUFE',
            'SS-CUDE': 'SS-CUDE', 
            'SS-SinAporte': 'SS-SinAporte',
            'SS-POS': 'SS-POS',
            'SS-SNum': 'SS-SNum',
            'SS-Recaudo': 'SS-Recaudo',
            'SS-Reporte': 'SS-Reporte',
            'Estandar': '10',
            'AIU': '09',
            'Mandatos': '11'
        }
        
        return fe_mapping.get(value, value)
    
    def _map_invoice_type(self, value):
        """Mapea el tipo de factura"""
        if not value:
            return 'insurance'
        
        value = str(value).strip().lower()
        
        type_mapping = {
            'aseguradora': 'insurance',
            'copago': 'copay',
            'regular': 'regular',
            'particular': 'regular',
            'ss-cufe': 'insurance',
            'ss-sinaporte': 'insurance',
            'ss-cude': 'insurance'
        }
        
        return type_mapping.get(value, 'insurance')
    
    def _find_city(self, city_name):
        """Busca ciudad por nombre"""
        if not city_name:
            return False
            
        return self.env['res.city'].search([
            ('name', 'ilike', city_name)
        ], limit=1)
    
    def _find_country(self, country_code):
        """Busca país por código numérico"""
        if not country_code:
            return False
        
        country_code = str(country_code).strip()
        
        country_mapping = {
            '170': 'CO',  # Colombia
            '840': 'US',  # Estados Unidos
            '484': 'MX',  # México
        }
        
        code = country_mapping.get(country_code)
        if code:
            return self.env['res.country'].search([('code', '=', code)], limit=1)
        
        return False
    
    def _validate_invoice_group(self, invoice_id, inv_data, row_nums):
        """Valida todos los datos de una factura antes de crearla"""
        errors = []
        
        # Validar encabezado
        header = inv_data['header']
        
        # Cliente obligatorio
        if not header.get('partner_id'):
            errors.append(f"Factura {invoice_id}: Cliente NIT/ID es obligatorio")
        else:
            partner = self._find_partner(header.get('partner_id'))
            if not partner:
                errors.append(f"Factura {invoice_id}: Cliente no encontrado: {header.get('partner_id')}")
        
        # Fecha de factura
        if header.get('invoice_date'):
            date_obj, error = self._validate_date(header.get('invoice_date'), 'Fecha Factura', row_nums[0])
            if error:
                errors.append(f"Factura {invoice_id}: {error}")
        
        # Fecha de vencimiento
        if header.get('invoice_date_due'):
            date_obj, error = self._validate_date(header.get('invoice_date_due'), 'Fecha Vencimiento', row_nums[0])
            if error:
                errors.append(f"Factura {invoice_id}: {error}")
        
        # Validar que tenga al menos una línea
        if not inv_data['lines']:
            errors.append(f"Factura {invoice_id}: No tiene líneas de detalle")
        
        # Validar cada línea
        for idx, line_data in enumerate(inv_data['lines']):
            line_errors = []
            
            # Producto obligatorio
            if not line_data.get('product_code'):
                line_errors.append("Código de producto es obligatorio")
            else:
                product = self._find_product(line_data.get('product_code'))
                if not product:
                    line_errors.append(f"Producto no encontrado: {line_data.get('product_code')}")
            
            # Cantidad y precio
            qty, error = self._validate_float(line_data.get('quantity'), 'Cantidad', row_nums[idx], required=True)
            if error:
                line_errors.append(error)
            elif qty <= 0:
                line_errors.append("La cantidad debe ser mayor a 0")

            price, error = self._validate_float(line_data.get('price_unit'), 'Precio Unitario', row_nums[idx], required=True)
            if error:
                line_errors.append(error)
            elif price < 0:
                line_errors.append("El precio no puede ser negativo")
            
            # Validar paciente si existe
            # if line_data.get('patient_document'):
            #     patient = self._find_or_create_patient(line_data)
            #     if not patient and not self.create_missing_patients:
            #         line_errors.append(f"Paciente no encontrado: {line_data.get('patient_document')}")
            
            # Fecha de atención
            if line_data.get('fecha_atencion'):
                date_obj, error = self._validate_date(line_data.get('fecha_atencion'), 'Fecha Atención', row_nums[idx])
                if error:
                    line_errors.append(error)
            
            # Agregar errores de línea
            if line_errors:
                errors.append(f"Factura {invoice_id}, Línea {idx+1} (Fila {row_nums[idx]}): " + "; ".join(line_errors))
        
        return errors
    
    def action_import(self):
        """Importa las facturas directamente sin previsualización"""
        self.ensure_one()
        
        if not self.file:
            raise UserError(_("Por favor, cargue un archivo Excel primero."))
        
        try:
            import base64, io
            data = base64.b64decode(self.file)
            
            # Leer archivo Excel
            try:
                import openpyxl
                book = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
                sheet = book.active
                
                headers = []
                for cell in next(sheet.iter_rows()):
                    headers.append(cell.value or "")
                
                rows = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    row_values = []
                    for cell in row:
                        row_values.append(cell.value)
                    if any(row_values):
                        rows.append((row_idx, row_values))
                        
            except:
                import xlrd
                book = xlrd.open_workbook(file_contents=data)
                sheet = book.sheet_by_index(0)
                
                headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                
                rows = []
                for row_idx in range(1, sheet.nrows):
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        if sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                            cell_value = xlrd.xldate.xldate_as_datetime(cell_value, book.datemode).date().isoformat()
                        row_values.append(cell_value)
                    if any(row_values):
                        rows.append((row_idx + 1, row_values))
            
            # Mapear columnas
            field_mapping = self._get_field_mapping()
            detected_columns = {}
            
            for idx, header in enumerate(headers):
                if header in field_mapping:
                    detected_columns[field_mapping[header]] = idx
            
            # Agrupar datos por número de factura
            invoices_data = {}
            
            for row_num, row_values in rows:
                id_temp = None
                if 'id_temp' in detected_columns:
                    id_temp = row_values[detected_columns['id_temp']]
                    
                if not id_temp:
                    continue
                    
                if id_temp not in invoices_data:
                    invoices_data[id_temp] = {
                        'header': {},
                        'lines': [],
                        'row_nums': []
                    }
                
                invoices_data[id_temp]['row_nums'].append(row_num)
                
                # Extraer campos de encabezado
                header_fields = [
                    'partner_id', 'invoice_date', 'invoice_date_due', 'journal_id',
                    'invoice_payment_term_id', 'fe_operation_type', 'invoice_type',
                    'date_start', 'date_end', 'health_payment_mode_id',
                    'health_coverage_plan_id', 'original_invoice_id',
                    'health_mipres_number', 'health_mipres_delivery_number', 'contract_id',
                    'x_ingreso'
                ]
                
                for field in header_fields:
                    if field in detected_columns and row_values[detected_columns[field]]:
                        if field not in invoices_data[id_temp]['header']:
                            invoices_data[id_temp]['header'][field] = row_values[detected_columns[field]]
                
                # Extraer campos de línea
                line_data = {}
                for field, idx in detected_columns.items():
                    if field not in header_fields and field != 'id_temp':
                        line_data[field] = row_values[idx]
                        
                if any(line_data.values()):
                    invoices_data[id_temp]['lines'].append(line_data)
            
            # Validar todas las facturas primero
            all_errors = []
            valid_invoices = {}
            
            for invoice_id, inv_data in invoices_data.items():
                errors = self._validate_invoice_group(invoice_id, inv_data, inv_data['row_nums'])
                if errors:
                    all_errors.extend(errors)
                else:
                    valid_invoices[invoice_id] = inv_data
            
            # Crear facturas válidas
            created_invoices = self.env['account.move']
            invoices_created = 0
            lines_created = 0
            
            for invoice_id, inv_data in valid_invoices.items():
                try:
                    # Preparar valores de factura
                    invoice_vals = self._prepare_invoice_vals(inv_data, invoice_id)
                    invoice = self.env['account.move'].create(invoice_vals)
                    created_invoices += invoice
                    invoices_created += 1
                    
                    # Crear líneas
                    for line_data in inv_data['lines']:
                        line_vals = self._prepare_invoice_line_vals(line_data, invoice.id)
                        self.env['account.move.line'].create(line_vals)
                        lines_created += 1
                        
                except Exception as e:
                    # Si falla la creación, eliminar la factura parcial si existe
                    if 'invoice' in locals() and invoice.exists():
                        invoice.unlink()
                    all_errors.append(f"Factura {invoice_id}: Error al crear - {str(e)}")
            
            # Actualizar estadísticas
            self.write({
                'total_invoices_processed': invoices_created,
                'total_lines_processed': lines_created,
                'total_invoices_failed': len(invoices_data) - invoices_created,
                'total_lines_failed': sum(len(inv['lines']) for inv in invoices_data.values()) - lines_created,
                'has_errors': bool(all_errors),
                'error_count': len(all_errors)
            })
            
            # Generar archivo de errores si hay
            if all_errors:
                import datetime
                error_content = f"REPORTE DE ERRORES DE IMPORTACIÓN\n"
                error_content += f"Fecha: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                error_content += f"Archivo: {self.file_name}\n"
                error_content += f"{'=' * 80}\n\n"
                error_content += f"Total de errores encontrados: {len(all_errors)}\n\n"
                error_content += f"DETALLE DE ERRORES:\n"
                error_content += f"{'-' * 80}\n"
                
                for error in all_errors:
                    error_content += f"• {error}\n"
                
                error_content += f"\n{'=' * 80}\n"
                error_content += f"RESUMEN:\n"
                error_content += f"- Facturas procesadas exitosamente: {invoices_created}\n"
                error_content += f"- Líneas procesadas exitosamente: {lines_created}\n"
                error_content += f"- Facturas con errores: {self.total_invoices_failed}\n"
                error_content += f"- Líneas con errores: {self.total_lines_failed}\n"
                
                filename = f"errores_importacion_{fields.Date.today()}.txt"
                self.write({
                    'error_log': base64.encodebytes(error_content.encode()),
                    'error_filename': filename
                })
            
            # Mensaje de resultado
            if created_invoices:
                message = f"""
                <h3>Importación completada</h3>
                <p><strong>Resumen del proceso:</strong></p>
                <ul>
                    <li>✅ Facturas creadas: {invoices_created}</li>
                    <li>✅ Líneas creadas: {lines_created}</li>
                    <li>❌ Facturas con errores: {self.total_invoices_failed}</li>
                    <li>❌ Líneas con errores: {self.total_lines_failed}</li>
                </ul>
                """
                
                if all_errors:
                    message += f"<p>⚠️ Se encontraron {len(all_errors)} errores. Descargue el archivo de errores para más detalles.</p>"
                
                return {
                    'name': _('Facturas Importadas'),
                    'type': 'ir.actions.act_window',
                    'res_model': 'account.move',
                    'view_mode': 'tree,form',
                    'domain': [('id', 'in', created_invoices.ids)],
                    'context': {
                        'search_default_group_by_partner': 1,
                        'import_message': message
                    }
                }
            # else:
            #     raise UserError(_(
            #         f"No se pudieron crear facturas.\n\n"
            #         f"Total de errores encontrados: {len(all_errors)}\n"
            #         f"Descargue el archivo de errores para ver el detalle."
            #     ))
                
        except Exception as e:
            raise UserError(_(f"Error al procesar el archivo: {str(e)}"))
    
    def _prepare_invoice_vals(self, inv_data, invoice_id_temp):
        """Prepara valores para crear la factura"""
        header = inv_data['header']
        
        # Cliente
        partner = self._find_partner(header.get('partner_id'))
        if not partner:
            raise ValidationError(_(f"No se encontró el cliente: {header.get('partner_id')}"))
        
        # Fechas
        invoice_date, _ = self._validate_date(header.get('invoice_date'), 'invoice_date', 0)
        invoice_date = invoice_date or self.invoice_date
        
        date_due, _ = self._validate_date(header.get('invoice_date_due'), 'invoice_date_due', 0) 
        date_start, _ = self._validate_date(header.get('date_start'), 'date_start', 0)
        date_start = date_start or invoice_date
        date_end, _ = self._validate_date(header.get('date_end'), 'date_end', 0)
        date_end = date_end or invoice_date
        
        # Diario y términos
        journal = self._find_journal(header.get('journal_id'))
        payment_term = self._find_payment_term(header.get('invoice_payment_term_id'))
        
        # Generar número de factura
        invoice_name = f"{journal.code}{invoice_id_temp}"
        
        # Mapear campos de salud
        fe_operation_type = self._map_fe_operation_type(header.get('invoice_type'))
        invoice_type = self._map_invoice_type(header.get('fe_operation_type'))
        
        # Buscar entidades de salud
        health_payment_mode = self._find_health_payment_mode(header.get('health_payment_mode_id'))
        health_coverage_plan = self._find_health_coverage_plan(header.get('health_coverage_plan_id'))
        health_collection_concept = self._find_health_collection_concept(header.get('health_payment_mode_id'))
        contract = self._find_contract(header.get('contract_id'))
        
        return {
            'partner_id': partner.id,
            'invoice_date': invoice_date,
            'invoice_date_due': date_due,
            'journal_id': journal.id,
            'move_type': 'out_invoice',
            'company_id': self.company_id.id,
            'health_provider_code': self.company_id.partner_id.ref,
            'invoice_payment_term_id': payment_term.id if payment_term else False,
            'invoice_origin': str(self.file_name).split('.')[0], #+ '-' + str(invoice_id_temp),
            'name': invoice_name,
            'contract_id': contract.id if contract else False,
            'x_ingreso': header.get('x_ingreso', ''),
            # Campos de salud
            'is_health_sector': True,
            'fe_operation_type': fe_operation_type,
            'invoice_type': invoice_type,
            'date_start': date_start,
            'date_end': date_end,
            'health_payment_mode_id': health_payment_mode.id if health_payment_mode else False,
            'health_coverage_plan_id': health_coverage_plan.id if health_coverage_plan else False,
            'health_collection_concept_id': health_collection_concept.id if health_collection_concept else False,
            'health_mipres_number': header.get('health_mipres_number', ''),
            'health_mipres_delivery_number': header.get('health_mipres_delivery_number', ''),
        }
    
    def _prepare_invoice_line_vals(self, line_data, move_id):
        """Prepara valores para crear líneas de factura"""
        # Producto
        product = self._find_product(line_data.get('product_code'))
        if not product:
            raise ValidationError(_(f"No se encontró el producto: {line_data.get('product_code')}"))
        
        # Paciente
        #patient = self._find_or_create_patient(line_data)
        
        # Valores numéricos
        quantity, _ = self._validate_float(line_data.get('quantity'), 'quantity', 0)
        quantity = quantity or 1.0
        price_unit, _ = self._validate_float(line_data.get('price_unit'), 'price_unit', 0)
        
        # Fechas
        fecha_atencion, _ = self._validate_date(line_data.get('fecha_atencion'), 'fecha_atencion', 0)
        if fecha_atencion:
            from datetime import datetime
            fecha_atencion = datetime.combine(fecha_atencion, datetime.min.time().replace(hour=8))
        
        patient_birth_date, _ = self._validate_date(line_data.get('patient_birth_date'), 'patient_birth_date', 0)
        
        # Buscar ciudad y país
        patient_city = self._find_city(line_data.get('patient_city_id'))
        patient_country = self._find_country(line_data.get('patient_country_id'))
        
        # Mapear tipo de usuario
        user_type = str(line_data.get('patient_user_type', '')).lower()
        user_type_mapping = {
            'contributivo cotizante': '01',
            'contributivo beneficiario': '02',
            'subsidiado': '04',
            'particular': '08',
        }
        mapped_user_type = user_type_mapping.get(user_type, '04')
        
        # Cuenta contable
        account = product.categ_id.property_account_income_categ_id
        if not account and self.default_account_id:
            account = self.default_account_id
        
        return {
            'product_id': product.id,
            'name': line_data.get('name') or product.name,
            'move_id': move_id,
            'quantity': quantity,
            'price_unit': price_unit,
            'product_uom_id': product.uom_id.id,
            'account_id': account.id if account else False,
            
            # Paciente
            #'patient_id': patient.id if patient else False,
            'patient_document': str(line_data.get('patient_document', '')),
            'patient_name': line_data.get('patient_name', ''),
            'patient_city_id': patient_city.id if patient_city else False,
            'patient_country_id': patient_country.id if patient_country else False,
            'patient_nationality': line_data.get('patient_nationality', ''),
            'patient_gender': str(line_data.get('patient_gender', '')),
            'patient_user_type': mapped_user_type,
            'patient_zone': str(line_data.get('patient_zone', '')).lower(),
            'patient_birth_date': patient_birth_date,
            'patient_doc_type': str(line_data.get('patient_doc_type', '')).upper(),
            
            # Otros campos
            'diagnostico_principal': line_data.get('diagnostico_principal', ''),
            'autorizacion': str(line_data.get('autorizacion', '')),
            'fecha_atencion': fecha_atencion,
            'fecha_procedimiento': fecha_atencion,
        }
    
    def action_new_import(self):
        """Reinicia el wizard para una nueva importación"""
        self.ensure_one()
        
        # Limpiar todos los campos
        self.write({
            'file': False,
            'file_name': False,
            'has_errors': False,
            'error_count': 0,
            'error_log': False,
            'error_filename': False,
            'total_invoices_processed': 0,
            'total_lines_processed': 0,
            'total_invoices_failed': 0,
            'total_lines_failed': 0,
        })
        
        # Retornar la misma vista
        return {
            'name': _('Importar Facturas Sector Salud'),
            'type': 'ir.actions.act_window',
            'res_model': 'health.invoice.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def update_patient_from_invoice(self, invoice_number, patient_fields):
        """
        Actualiza información del paciente basado en el número de factura
        
        :param invoice_number: Número de la factura
        :param patient_fields: Diccionario con los campos a actualizar del paciente
        :return: Número de pacientes actualizados
        """
        # Buscar factura por número
        invoice = self.env['account.move'].search([
            ('name', '=', invoice_number),
            ('move_type', '=', 'out_invoice')
        ], limit=1)
        
        if not invoice:
            raise UserError(_(f"No se encontró la factura con número: {invoice_number}"))
        
        # Obtener pacientes únicos de las líneas
        patients = invoice.invoice_line_ids.mapped('patient_id')
        patients = patients.filtered(lambda p: p)  # Filtrar pacientes vacíos
        
        if not patients:
            raise UserError(_(f"La factura {invoice_number} no tiene pacientes asociados"))
        
        # Actualizar cada paciente
        updated_count = 0
        for patient in patients:
            try:
                # Validar campos antes de actualizar
                update_vals = {}
                
                if 'name' in patient_fields:
                    update_vals['name'] = patient_fields['name']
                
                if 'date_of_birth' in patient_fields:
                    date_obj, error = self._validate_date(patient_fields['date_of_birth'], 'date_of_birth', 0)
                    if date_obj:
                        update_vals['date_of_birth'] = date_obj
                
                if 'email' in patient_fields:
                    update_vals['email'] = patient_fields['email']
                
                if 'phone' in patient_fields:
                    update_vals['phone'] = patient_fields['phone']
                
                if 'mobile' in patient_fields:
                    update_vals['mobile'] = patient_fields['mobile']
                
                # Agregar otros campos según sea necesario
                
                if update_vals:
                    patient.write(update_vals)
                    updated_count += 1
                    
            except Exception as e:
                _logger.error(f"Error actualizando paciente {patient.name}: {str(e)}")
        
        return updated_count